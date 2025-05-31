from flask import Flask, render_template, request, jsonify, redirect, url_for
from Phidget22.Phidget import *
from Phidget22.Devices.DigitalInput import *
import pyvisa
import os
import json
import serial
import threading
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
import requests
import atexit
import serial.tools.list_ports
import webview

app = Flask(__name__)
rm = pyvisa.ResourceManager('@py')

arduino_serial = None
psw_device = None
device = None
stop_flag = False
settings_file = "settings.json"
latest_cycle_data = []
latest_log_text = ""
latest_status = "Waiting for result..."
cycle_running = False
cycle_lock = threading.Lock()
current_cycle = 0 

# Load or create default settings
def load_settings():
    default_settings = {
        "voltage": "24.00",
        "current": "3.00",
        "cycles": 10,
        "delay": 1.0,
        "username": "Pawan",
        "password": "Pawan@123",
        "excel_path": "D/PSW84"
    }

    if not os.path.exists(settings_file):
        with open(settings_file, 'w') as f:
            json.dump(default_settings, f)
        return default_settings
    else:
        with open(settings_file, 'r') as f:
            settings = json.load(f)

        for key in default_settings:
            if key not in settings:
                settings[key] = default_settings[key]

        with open(settings_file, 'w') as f:
            json.dump(settings, f)

        return settings

@app.route("/")
def dashboard():
    return render_template("dashboard.html")

@app.route('/get_settings')
def get_settings():
    try:
        with open(settings_file, "r") as f:
            settings = json.load(f)
        return jsonify({
            "voltage": settings.get("voltage", "0.00"),
            "current": settings.get("current", "0.00"),
            "cycles": int(settings.get("cycles", 0)),
            "delay": float(settings.get("delay", 1.0))  # <-- add this line
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/settings", methods=["GET", "POST"])
def settings_page():
    if request.method == "POST":
        try:
            voltage = request.form.get("voltage", "0.00").strip()
            current = request.form.get("current", "0.00").strip()
            cycles = int(request.form.get("cycles", 0))
            delay = round(float(request.form.get("delay", 0)), 2)

            # Validate float format
            float(voltage)
            float(current)

            if os.path.exists(settings_file):
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
            else:
                settings = {}

            settings["voltage"] = voltage
            settings["current"] = current
            settings["cycles"] = cycles
            settings["delay"] = delay

            with open(settings_file, 'w') as f:
                json.dump(settings, f)

            return redirect(url_for("settings_page"))
        except Exception as e:
            return f"Error in saving settings: {str(e)}", 400

    else:
        if os.path.exists(settings_file):
            with open(settings_file, 'r') as f:
                settings = json.load(f)
        else:
            settings = {
                "voltage": "24.00",
                "current": "3.00",
                "cycles": 5,
                "delay": 1.0,
                "username": "pawan",
                "password": "Pawan@123",
                "excel_path": "."
            }

        return render_template("settings.html", settings=settings)

@app.route("/verify_login", methods=["POST"])
def verify_login():
    data = request.json
    entered_username = data.get("username", "").lower()
    entered_password = data.get("password", "")
    settings = load_settings()

    stored_username = settings.get("username", "").lower()
    stored_password = settings.get("password", "")

    return jsonify({"success": entered_username == stored_username and entered_password == stored_password})

@app.route("/cycle_data")
def reading_data():
    return render_template("cycle_data.html", cycle_data=latest_cycle_data)

@app.route("/connect")
def connect():
    global psw_device, device
    try:
        if device is not None:
            return jsonify({"status": "Already Connected", "message": "Device already connected."})

        available_ports = rm.list_resources()
        serial_ports = [port for port in available_ports if "ASRL" in port]

        if not serial_ports:
            return jsonify({"status": "Error", "message": "No COM ports found"})

        psw_device = rm.open_resource(serial_ports[0])
        psw_device.baud_rate = 9600
        psw_device.data_bits = 8
        psw_device.parity = pyvisa.constants.Parity.none
        psw_device.stop_bits = pyvisa.constants.StopBits.one
        psw_device.timeout = 500

        idn = psw_device.query("*IDN?")
        device = psw_device

        return jsonify({"status": "Connected", "message": f"{idn.strip()}"})
    except Exception as e:
        return jsonify({"status": "Error", "message": f"Failed to connect: {str(e)}"})

@app.route("/disconnect")
def disconnect():
    global psw_device, device
    try:
        if psw_device:
            psw_device.close()
        psw_device = None
        device = None
        return jsonify({"status": "Disconnected", "message": "Device disconnected successfully."})
    except Exception as e:
        return jsonify({"status": "Error", "message": f"Disconnection error: {str(e)}"})

@app.route("/status")
def status():
    return jsonify({"connected": device is not None})

@app.route("/start_loop", methods=["POST"])
def start_loop():
    global device, stop_flag, latest_cycle_data, latest_log_text, latest_status, cycle_running, current_cycle

    with cycle_lock:
        if cycle_running:
            return jsonify({"status": "Running", "message": "Cycle already running"})
        cycle_running = True
        current_cycle = 0  # Reset cycle count at start

    settings = load_settings()
    voltage = round(float(settings["voltage"]), 2)
    current = round(float(settings["current"]), 2)
    cycles = int(settings["cycles"])
    delay = round(float(settings["delay"]), 2)

    if not device:
        cycle_running = False
        return jsonify({"status": "Error", "message": "Device not connected"})

    result_data = []
    stop_flag = False
    latest_log_text = ""
    latest_status = "Cycle Running..."

    def run_cycles():
        nonlocal result_data
        global stop_flag, latest_cycle_data, latest_log_text, latest_status, cycle_running, current_cycle

        job_start_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            for i in range(1, cycles + 1):
                if stop_flag:
                    break

                current_cycle = i  # Update current cycle here

                try:
                    device.write("OUTP OFF")
                    device.write(f"APPL {voltage},{current}")
                    device.write("OUTP ON")
                    time.sleep(delay)

                    measured_voltage = float(device.query("MEAS:VOLT?"))
                    measured_current = float(device.query("MEAS:CURR?"))

                    device.write("OUTP OFF")
                    time.sleep(delay)
                    device.write("OUTP ON")

                    status = "OK"
                    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    result_data.append({
                        "cycle": i,
                        "voltage": round(measured_voltage, 2),
                        "current": round(measured_current, 2),
                        "status": status,
                        "timestamp": timestamp
                    })

                    latest_log_text += f"[{i}/{cycles}] Cycle {i}: V={measured_voltage:.2f}V, I={measured_current:.2f}A, Status={status}, Time={timestamp}\n"

                except Exception as e:
                    error_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    latest_log_text += f"[{i}/{cycles}] Cycle {i} Error: {str(e)} at {error_time}\n"

                    result_data.append({
                        "cycle": i,
                        "voltage": 0,
                        "current": 0,
                        "status": f"Error: {str(e)}",
                        "timestamp": error_time
                    })

            device.write("OUTP OFF")

            if stop_flag or len(result_data) < cycles:
                latest_status = "Stopped early - Job Not OK"
            else:
                latest_status = "Job OK"

            save_to_excel(result_data, job_start_time, voltage)
            latest_cycle_data = result_data

        except Exception as err:
            latest_log_text += f"Critical error during job: {str(err)}\n"
            latest_status = "Job Not OK"
        finally:
            cycle_running = False
            current_cycle = cycles if not stop_flag else current_cycle

    threading.Thread(target=run_cycles).start()
    return jsonify({"status": "Started", "message": "Cycle loop started"})

# New endpoint to provide progress to frontend
@app.route("/cycle_progress")
def cycle_progress():
    settings = load_settings()
    total_cycles = int(settings.get("cycles", 0))
    global current_cycle

    return jsonify({
        "current": current_cycle,
        "total": total_cycles
    })

@app.route("/stop_loop", methods=["POST"])
def stop_loop():
    global stop_flag
    stop_flag = True
    return jsonify({"status": "Stopped", "message": "Cycle loop stopped"})


@app.route("/cycle_status")
def cycle_status():
    return jsonify({
        "logs": latest_log_text,
        "status": latest_status
    })

def save_to_excel(data, start_time, voltage):
    settings = load_settings()
    user_path = settings.get("excel_path", "").strip()

    # Get current year and month
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")  # Two-digit month

    # Construct folder hierarchy: base/year/month
    if user_path:
        base_dir = user_path
    else:
        base_dir = "."

    save_dir = os.path.join(base_dir, year, month)

    try:
        os.makedirs(save_dir, exist_ok=True)
    except Exception as e:
        print(f"[WARNING] Cannot create directory '{save_dir}', saving to current directory instead. Error: {e}")
        save_dir = "."

    current_date = now.strftime("%Y-%m-%d")
    logs_excel_path = os.path.join(save_dir, f"logs_{current_date}.xlsx")
    results_excel_path = os.path.join(save_dir, f"Cycle_Result_{current_date}.xlsx")

    # Save cycle logs
    if not os.path.exists(logs_excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Log Start Time", start_time])
        ws.append(["Cycle", "Voltage", "Current", "Status", "Time"])
    else:
        wb = load_workbook(logs_excel_path)
        ws = wb.active
        ws.append([])  # Blank line for separation
        ws.append(["Log Start Time", start_time])
        ws.append(["Cycle", "Voltage", "Current", "Status", "Time"])

    for entry in data:
        ws.append([entry["cycle"], entry["voltage"], entry["current"], entry["status"], entry["timestamp"]])
    wb.save(logs_excel_path)

    # Save job summary result
    if not os.path.exists(results_excel_path):
        result_wb = Workbook()
        result_ws = result_wb.active
        result_ws.append(["Job Start", "Voltage", "Total Cycles", "Status"])
    else:
        result_wb = load_workbook(results_excel_path)
        result_ws = result_wb.active

    all_ok = all(entry["status"] == "OK" for entry in data)
    result_ws.append([start_time, voltage, len(data), "OK" if all_ok else "Not OK"])
    result_wb.save(results_excel_path)

    print(f"[INFO] Excel files saved to: {save_dir}")

# Arduino listener thread for button press
# PHIDGET: Replace Arduino serial button listener with Phidget Digital Input


start_button = None
stop_button = None
phidget_thread = None

START_URL = "http://127.0.0.1:5001/start_loop"
STOP_URL = "http://127.0.0.1:5001/stop_loop"

# Button press handler
def on_button_pressed(self, state):
    if not state:
        return  # Ignore release events

    channel = self.getChannel()

    if channel == 0:
        print("[BUTTON] Start button pressed.")
        try:
            response = requests.post(START_URL)
            print(f"[START] {response.status_code} - {response.text}")
        except Exception as e:
            print(f"[ERROR] Failed to send start POST: {e}")
    elif channel == 1:
        print("[BUTTON] Stop button pressed.")
        try:
            response = requests.post(STOP_URL)
            print(f"[STOP] {response.status_code} - {response.text}")
        except Exception as e:
            print(f"[ERROR] Failed to send stop POST: {e}")

# Listen for button presses in background
def listen_to_phidget_buttons():
    global start_button, stop_button

    try:
        print("[INFO] Connecting to Phidget buttons...")

        # Setup Start button (Channel 0)
        start_button = DigitalInput()
        start_button.setChannel(0)
        start_button.setOnStateChangeHandler(on_button_pressed)
        start_button.openWaitForAttachment(5000)

        # Setup Stop button (Channel 1)
        stop_button = DigitalInput()
        stop_button.setChannel(1)
        stop_button.setOnStateChangeHandler(on_button_pressed)
        stop_button.openWaitForAttachment(5000)

        print("[INFO] Both Phidget buttons attached and listening.")

        while True:
            time.sleep(0.1)

    except PhidgetException as e:
        print(f"[ERROR] Phidget Exception: {e}")
    finally:
        if start_button:
            try:
                start_button.close()
                print("[INFO] Start button closed.")
            except Exception as e:
                print(f"[WARN] Error closing start button: {e}")
        if stop_button:
            try:
                stop_button.close()
                print("[INFO] Stop button closed.")
            except Exception as e:
                print(f"[WARN] Error closing stop button: {e}")

# Thread launcher
def run_in_thread():
    global phidget_thread
    phidget_thread = threading.Thread(target=listen_to_phidget_buttons, daemon=True)
    phidget_thread.start()

# Cleanup on exit
def cleanup():
    global start_button, stop_button
    if start_button:
        try: start_button.close()
        except: pass
    if stop_button:
        try: stop_button.close()
        except: pass

atexit.register(cleanup)


if __name__ == "__main__":
    # Start Arduino button listener in the background.
    run_in_thread()
    
    # Function to run Flask in a separate thread.
    def run_flask():
        # Disable the reloader to prevent signal handling in a non-main thread.
        app.run(debug=True, use_reloader=False, port=5001)

    
    # Start Flask in a daemon thread.
    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()
    
    # Optionally sleep for a couple seconds to allow the server to start.
    time.sleep(2)
    
    # Launch the PyWebView window pointing to the Flask app.
    webview.create_window("Mahindra PSW Controller", "http://127.0.0.1:5001", width=1200, height=800)
    
    # This call starts the webview event loop.
    webview.start()